"""Build a single Excel workbook comparing GAN vs Diffusion fusion metrics.

Reads the per-pair GAN summary rows (outputs/models/gan/metrics/<pair>_metrics.csv)
and the diffusion summary (outputs/models/diffusion/metrics/diffusion_metrics_summary.csv),
then writes a styled comparison workbook to outputs/comparison/.

All metric cells keep the "mean±std" string format. The model with the better
mean (higher-is-better for every metric here) is bold per metric per pair.
"""

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]

PAIRS = ["ct_mri", "pet_mri", "spect_mri"]
PAIR_LABELS = {"ct_mri": "CT-MRI", "pet_mri": "PET-MRI", "spect_mri": "SPECT-MRI"}
METRICS = ["SSIM", "PSNR", "MI", "EN", "CC", "FMI", "SF", "AG"]
DIFFUSION_VARIANT = "diffusion_fused_original"  # primary final fused output

# Styling
HEADER_FILL = PatternFill("solid", fgColor="305496")
PAIR_FILL = PatternFill("solid", fgColor="D9E1F2")
GAN_FILL = PatternFill("solid", fgColor="FCE4D6")
DIFF_FILL = PatternFill("solid", fgColor="E2EFDA")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WIN_FONT = Font(bold=True, color="006100")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")


def mean_of(text):
    """Extract the numeric mean from a 'mean±std' (or 'mean +/- std') string."""
    s = str(text or "").strip()
    for sep in ("±", "+/-"):
        if sep in s:
            s = s.split(sep, 1)[0].strip()
            break
    try:
        return float(s)
    except ValueError:
        return float("nan")


def read_gan(metrics_dir):
    """Return {pair: {metric: 'mean±std'}} from the per-pair GAN summary rows."""
    data = {}
    for pair in PAIRS:
        path = metrics_dir / f"{pair}_metrics.csv"
        if not path.exists():
            print(f"[warn] missing GAN file: {path}")
            continue
        with path.open(encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        summary = rows[-1]  # last row is the mean±std summary
        data[pair] = {m: summary.get(m, "") for m in METRICS}
    return data


def read_diffusion(metrics_dir, variant):
    """Return {pair: {metric: 'mean±std'}} from the diffusion summary CSV."""
    path = metrics_dir / "diffusion_metrics_summary.csv"
    data = {}
    if not path.exists():
        print(f"[warn] missing diffusion file: {path}")
        return data
    with path.open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            if row.get("variant") == variant:
                data[row["pair"]] = {m: row.get(m, "") for m in METRICS}
    return data


def style_header(ws, row, headers):
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def autosize(ws, widths):
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def build_comparison_sheet(ws, gan, diff):
    ws.title = "Comparison"
    ws.cell(row=1, column=1, value="GAN vs Diffusion — Fusion Metrics (mean±std)").font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(METRICS) + 2)

    headers = ["Pair", "Model", *METRICS]
    style_header(ws, 2, headers)

    r = 3
    for pair in PAIRS:
        g = gan.get(pair, {})
        d = diff.get(pair, {})
        for model, src, fill in (("GAN", g, GAN_FILL), ("Diffusion", d, DIFF_FILL)):
            ws.cell(row=r, column=1, value=PAIR_LABELS[pair]).border = BORDER
            ws.cell(row=r, column=1).fill = PAIR_FILL
            ws.cell(row=r, column=1).alignment = CENTER
            ws.cell(row=r, column=1).font = Font(bold=True)
            mcell = ws.cell(row=r, column=2, value=model)
            mcell.border = BORDER
            mcell.fill = fill
            mcell.alignment = CENTER
            for i, metric in enumerate(METRICS):
                cell = ws.cell(row=r, column=3 + i, value=src.get(metric, ""))
                cell.alignment = CENTER
                cell.border = BORDER
                # bold the winner (higher mean is better for all these metrics)
                gm, dm = mean_of(g.get(metric)), mean_of(d.get(metric))
                if gm == gm and dm == dm:  # both not NaN
                    winner = "GAN" if gm > dm else ("Diffusion" if dm > gm else None)
                    if winner == model:
                        cell.font = WIN_FONT
            r += 1
        # merge the pair label across its two model rows
        ws.merge_cells(start_row=r - 2, start_column=1, end_row=r - 1, end_column=1)

    autosize(ws, [12, 12] + [13] * len(METRICS))
    ws.freeze_panes = "C3"

    note = r + 1
    ws.cell(row=note, column=1,
            value=f"GAN = canonical test.py metrics. Diffusion = '{DIFFUSION_VARIANT}'. "
                  f"Bold green = better mean per metric per pair (higher is better).").font = Font(italic=True, size=9, color="808080")


def build_single_model_sheet(ws, title, data):
    ws.title = title
    headers = ["Pair", *METRICS]
    style_header(ws, 1, headers)
    for r, pair in enumerate(PAIRS, start=2):
        ws.cell(row=r, column=1, value=PAIR_LABELS[pair]).font = Font(bold=True)
        ws.cell(row=r, column=1).fill = PAIR_FILL
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=1).border = BORDER
        for i, metric in enumerate(METRICS):
            cell = ws.cell(row=r, column=2 + i, value=data.get(pair, {}).get(metric, ""))
            cell.alignment = CENTER
            cell.border = BORDER
    autosize(ws, [12] + [13] * len(METRICS))


def build_diffusion_all_variants_sheet(ws, metrics_dir):
    ws.title = "Diffusion (all variants)"
    path = metrics_dir / "diffusion_metrics_summary.csv"
    if not path.exists():
        return
    with path.open(encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        return
    style_header(ws, 1, rows[0])
    for r, row in enumerate(rows[1:], start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = CENTER
            cell.border = BORDER
    autosize(ws, [11, 26] + [13] * (len(rows[0]) - 2))


def main():
    parser = argparse.ArgumentParser(description="Build GAN-vs-Diffusion comparison Excel.")
    parser.add_argument("--gan-metrics", default=str(PROJECT_ROOT / "outputs" / "models" / "gan" / "metrics"))
    parser.add_argument("--diffusion-metrics", default=str(PROJECT_ROOT / "outputs" / "models" / "diffusion" / "metrics"))
    parser.add_argument("--output", default=str(PROJECT_ROOT / "outputs" / "comparison" / "model_metrics_comparison.xlsx"))
    parser.add_argument("--diffusion-variant", default=DIFFUSION_VARIANT)
    args = parser.parse_args()

    gan_dir = Path(args.gan_metrics)
    diff_dir = Path(args.diffusion_metrics)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    gan = read_gan(gan_dir)
    diff = read_diffusion(diff_dir, args.diffusion_variant)

    wb = Workbook()
    build_comparison_sheet(wb.active, gan, diff)
    build_single_model_sheet(wb.create_sheet("GAN"), "GAN", gan)
    build_single_model_sheet(wb.create_sheet("Diffusion"), "Diffusion", diff)
    build_diffusion_all_variants_sheet(wb.create_sheet("Diffusion (all variants)"), diff_dir)

    wb.save(out_path)
    print(f"Saved comparison workbook: {out_path}")
    for pair in PAIRS:
        print(f"  {PAIR_LABELS[pair]:9s}  GAN SSIM {gan.get(pair, {}).get('SSIM','-'):>12}  | Diffusion SSIM {diff.get(pair, {}).get('SSIM','-'):>12}")


if __name__ == "__main__":
    main()
